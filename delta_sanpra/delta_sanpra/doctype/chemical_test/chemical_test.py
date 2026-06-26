	# Copyright (c) 2026, Sanpra Software Solution and contributors
	# For license information, please see license.txt

import io
import frappe
from openpyxl import load_workbook
from frappe.utils import nowdate, nowtime
from frappe.model.document import Document
from frappe.utils import cint
import pdfplumber
import re

class ChemicalTest(Document):
	def get_chemical_value_color(self, row):
		value = float(row.value) if row.value not in (None, "") else None
		min_range = float(row.min_range) if row.min_range not in (None, "") else None
		max_range = float(row.max_range) if row.max_range not in (None, "") else None

		if value is None or min_range is None or max_range is None:
			return ""

		if value < min_range or value > max_range:
			return "#FF7A7A"

		return "#7CFF7C"

	@frappe.whitelist()
	def get_highlight_colors(self):
		tables = ["test_details_chemical"]
		colors = {}
		for table in tables:
			if not hasattr(self, table):
				continue
			for row in getattr(self, table):
				try:
					colors[row.name] = self.get_chemical_value_color(row)
				except (TypeError, ValueError):
					colors[row.name] = ""
		return colors
	#*************************************************************************************
	@frappe.whitelist()
	def get_minmax_range(self, test_method=None, parameter=None, material_specification=None):
		data = {}
		# Only if test_method provided
		if test_method:
			test_method_doc = frappe.get_doc("Test Method", test_method)
			for row in test_method_doc.chemical_details:
				if row.parameter == parameter:
					data["method_min_range"] = row.min_range
					data["method_max_range"] = row.max_range
					break
		# Material Specification always check
		if material_specification:
			material_spec_doc = frappe.get_doc("Item", material_specification)
			for row in material_spec_doc.custom_chemical_detail:
				if row.parameter == parameter:
					data["min_range"] = row.min_range
					data["max_range"] = row.max_range
					break
		return [data]


#**************************************************************************
	@frappe.whitelist()
	def set_ulr_counter(self):
		if self.ulr_no:
			return
     	# frappe.msgprint("Setting ULR Counter")
		company_name = "DELTAA METALLIX SOLUTIONS PRIVATE LIMITED"
		last = frappe.db.get_value("Company", company_name, "custom_ulr_counter") or 0
		count = int(last) + 1
		frappe.db.set_value("Company", company_name, "custom_ulr_counter", count)
		return count
#**************************************************************************
	@frappe.whitelist()
	def create_rate_chart_from_excel(self):
		if not self.upload_excel_file:
			return
		# Get File
		file_doc = frappe.get_doc("File", {"file_url": self.upload_excel_file})
		content = file_doc.get_content()
		workbook = load_workbook(io.BytesIO(content), data_only=True)
		sheet = workbook.active
		excel_data = {}
		# Read Excel Data
		for r in range(2, sheet.max_row + 1):
			parameter = sheet.cell(r, 1).value
			value = sheet.cell(r, 2).value
			if parameter:
				key = str(parameter).strip().lower()
				excel_data[key] = value
		# Update Child Table
		for row in self.test_details_chemical:
			if row.parameter:
				key = str(row.parameter).strip().lower()
				if key in excel_data:
					row.value = excel_data[key]
					# Status Calculation
					# if row.value and row.method_min_range and row.method_max_range:
					if row.value is not None and row.method_min_range is not None and row.method_max_range is not None:

						value = float(row.value)
						method_min_range = float(row.method_min_range)
						method_max_range = float(row.method_max_range)
						if method_min_range == 0 and method_max_range == 0:
							row.status = "NON NABL"
						elif value < method_min_range or value > method_max_range:
							row.status = "NON NABL"
						else:
							row.status = "NABL"
		self.save()
#***************************************************************************
	@frappe.whitelist()
	def read_pmi_excel(self):
		if not self.excel_attach:
			return
		# Get attached file
		file_doc = frappe.get_doc("File", {"file_url": self.excel_attach})
		content = file_doc.get_content()
		workbook = load_workbook(io.BytesIO(content), data_only=True)
		sheet = workbook.active
		# Clear old rows
		self.set("pmi_test_table", [])
		# Loop Excel rows (Assuming row 1 = header)
		for r in range(2, sheet.max_row + 1):
			element = sheet.cell(r, 1).value
			specified_range = sheet.cell(r, 2).value
			moc_result = sheet.cell(r, 3).value
			description = sheet.cell(r, 4).value
			if element:
				self.append("pmi_test_table", {
					"elements": element,
					"specified_range": specified_range,
					"mocresult": moc_result,
					"description": description
				})
		self.save()
#***************************************************************************	
	def on_update(self):
		self.update_sample_inward_printed()
		self.update_sample_inward_counters()

	def update_sample_inward_printed(self):
		# SAME code as above
		inward = self.inward_number
		if not inward:
			return
		
		test_doctypes = [
			"Physical Test",
			"Chemical Test", 
			"Corrosion Test",
			"Metallography Test",
			"Other Test"
		]
		
		printed_count = 0
		for dt in test_doctypes:
			tests = frappe.get_all(dt, filters={"inward_number": inward, "is_print": 1}, fields=["name"])
			printed_count += len(tests)
		
		frappe.db.set_value("Sample Inward", inward, "printed", printed_count)
		frappe.db.commit()
#*************************************************************************
	def update_sample_inward_counters(self):
		"""Dono counters update karo - NO RECURSION"""
		
		inward = self.inward_number
		if not inward:
			return
		
		# =========================================
		# 1. PRINTED COUNT
		# =========================================
		test_doctypes = [
			"Physical Test",
			"Chemical Test",
			"Corrosion Test",
			"Metallography Test",
			"Other Test"
		]
		
		printed_count = 0
		for dt in test_doctypes:
			count = frappe.db.count(dt, {
				"inward_number": inward,
				"is_print": 1
			})
			printed_count += count
		
		frappe.db.set_value("Sample Inward", inward, "printed", printed_count)
		
		# =========================================
		# 2. READY TO PRINT COUNT (NEW)
		# =========================================
		ready_count = 0
		for dt in test_doctypes:
			count = frappe.db.count(dt, {
				"inward_number": inward,
				"workflow_state": ["in", ["Approved"]]
			})
			ready_count += count
		
		frappe.db.set_value("Sample Inward", inward, "ready_to_print", ready_count)
		
		# Commit karo
		frappe.db.commit()