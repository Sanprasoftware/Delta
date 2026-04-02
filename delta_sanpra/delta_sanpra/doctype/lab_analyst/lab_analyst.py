# # Copyright (c) 2025, Sanpra Software Solution and contributors
# # For license information, please see license.txt
import io
import frappe
from openpyxl import load_workbook
from frappe.utils import nowdate, nowtime
from frappe.model.document import Document
from frappe.utils import cint
import pdfplumber
import re


class LabAnalyst(Document):
    def on_update(self):
        self.on_update_after_submit()
    def before_save(self):
        self.create_physical_details_from_excel()
        self.create_test_details_metallography()
        self.read_pdf_and_fill_table()
        # self.create_rate_chart_from_pdf()
     
        for row in self.test_details: 
            if not row.value:
                continue
            value = float(row.value)
            min_range = float(row.method_min_range or 0)
            max_range = float(row.method_max_range or 0)
            row.status = "NABL" if min_range < value < max_range else "NON NABL"

#*********************************************************************************************************

        # save document
        # self.save(ignore_permissions=True)

    # @frappe.whitelist()
    # def create_rate_chart_from_pdf(self):
    #     if not self.upload_pdf_file:
    #         return
    #     file_name = frappe.db.get_value("File", {"file_url": self.upload_pdf_file}, "name")
    #     file_doc = frappe.get_doc("File", file_name)
    #     file_content = file_doc.get_content()
    #     existing = {row.parameter: row for row in self.test_details if row.parameter}
    #     import pdfplumber, io
    #     with pdfplumber.open(io.BytesIO(file_content)) as pdf:
    #         for page in pdf.pages:
    #             text = page.extract_text()
    #             if not text:
    #                 continue
                # for line in text.split("\n"):
                #     if "Parameter" in line and "Value" in line:
                #         continue
                #     parts = line.split()
                #     if len(parts) < 2:
                #         continue
                #     param = " ".join(parts[:-1])
                #     val = parts[-1]

                #     if param in existing:
                #         existing[param].value = val
                #     else:
                #         self.append("test_details", {"parameter": param, "value": val})

    #********************************************************************************************************
    @frappe.whitelist()
    def create_rate_chart_from_excel(self):
        if not self.upload_excel_file:
            return
        file_doc = frappe.get_doc("File", {"file_url": self.upload_excel_file})
        sheet = load_workbook(io.BytesIO(file_doc.get_content()), data_only=True).active
        excel_data = {}
        for r in range(2, sheet.max_row + 1):
            p, v = sheet.cell(r, 1).value, sheet.cell(r, 2).value
            if p and v:
                excel_data[str(p).strip()] = str(v).strip()

        for row in self.test_details:
            param = str(row.parameter).strip() if row.parameter else ""
            if param in excel_data:
                row.value = excel_data[param]
#******************************************************************************************************** 
    def on_submit(self):
        self.get_child_table_id()
        self.on_update_after_submit()

    def get_child_table_id(self):
        inward_doc = frappe.get_doc("Sample Inward", self.inward_number)
        inward_number = inward_doc.name

        for row in inward_doc.test_on_sample:
            if row.name == self.child_table_id:
                frappe.db.set_value("Test On sample", row.name, "status", "Complete")
#********************************************************************************************************
    @frappe.whitelist()
    def result(self):
        for row in self.test_details:
            if not row.value:
                continue
            value = float(row.value)
            method_min_range = float(row.method_min_range)
            method_max_range = float(row.method_max_range)
            row.status = "NABL" if method_min_range < value < method_max_range else "NON NABL"            
# ***********************************************************************************************
    @frappe.whitelist()
    def create_physical_details_from_excel(self):
        if not self.excel_file:
            return
        file_doc = frappe.get_doc("File", {"file_url": self.excel_file})
        sheet = load_workbook(io.BytesIO(file_doc.get_content()), data_only=True).active
        self.test_details_physical = []
        headers = {}
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[cell.value.strip().lower()] = idx
        required_columns = ["parameter", "min range", "max range", "value"]
        for col in required_columns:
            if col not in headers:
                frappe.throw(f"Missing column in Excel: {col}")
        for r in range(2, sheet.max_row + 1):
            parameter = sheet.cell(r, headers["parameter"]).value
            min_range = sheet.cell(r, headers["min range"]).value
            max_range = sheet.cell(r, headers["max range"]).value
            value = sheet.cell(r, headers["value"]).value
            if parameter:
                self.append("test_details_physical", {
                    "parameter": str(parameter).strip(),
                    "min_range": float(min_range) if min_range not in (None, "") else None,
                    "max_range": float(max_range) if max_range not in (None, "") else None,
                    "value": float(value) if value not in (None, "") else None
                })
#***********************************************************************************************
    @frappe.whitelist()
    def get_highlight_colors(self):
        tables = ["test_details", "test_details_physical"]
        colors = {}
        for table in tables:
            if not hasattr(self, table):
                continue
            for row in getattr(self, table):
                try:
                    value = float(row.value) if row.value not in (None, "") else None
                    min_range = float(row.min_range) if row.min_range not in (None, "") else None
                    max_range = float(row.max_range) if row.max_range not in (None, "") else None

                    # Highlight if value < min_range OR value > max_range
                    if value is not None and min_range is not None and max_range is not None:
                        if value < min_range or value > max_range:
                            colors[row.name] = "#FF7A7A"     # RED
                        else:
                            colors[row.name] = "#7CFF7C"     # GREEN
                    else:
                        colors[row.name] = ""
                except:
                    colors[row.name] = ""
        return colors
#*********************************************************************************************************************
       
    @frappe.whitelist()
    def get_test_method(self, test_method):
        test_method_doc = frappe.get_doc("Test Method", test_method)
        parameters = [row.parameter for row in test_method_doc.chemical_details]
        # frappe.msgprint(str(parameters))
        return parameters

    @frappe.whitelist()
    def get_minmax_range(self, test_method, parameter, material_specification):
        data = {}
        test_method_doc = frappe.get_doc("Test Method", test_method)
        for row in test_method_doc.chemical_details:
            if row.parameter == parameter:
                data["method_min_range"] = row.min_range
                data["method_max_range"] = row.max_range
                break
        material_spec_doc = frappe.get_doc("Item", material_specification)
        for row in material_spec_doc.custom_chemical_detail:
            if row.parameter == parameter:
                data["min_range"] = row.min_range
                data["max_range"] = row.max_range
                break

        return [data] 
#**********************************************************************************************************
    @frappe.whitelist()
    def create_test_details_metallography(self):
        if not self.attach_excel_file:
            return
        file_doc = frappe.get_doc("File", {"file_url": self.attach_excel_file})
        sheet = load_workbook(io.BytesIO(file_doc.get_content()), data_only=True).active
        self.test_details_metallography = []
        headers = {cell.value.strip().lower(): idx for idx, cell in enumerate(sheet[1], start=1) if cell.value}
        if "parameter" not in headers or "value" not in headers:
            frappe.throw("Excel must contain 'Parameter' and 'Value' columns.")
        for r in range(2, sheet.max_row + 1):
            parameter = sheet.cell(r, headers["parameter"]).value
            value = sheet.cell(r, headers["value"]).value
            final_value = value
            try:
                final_value = float(value)
            except:
                if value is not None:
                    final_value = str(value).strip()
            self.append("test_details_metallography", {
                "parameter": str(parameter).strip(),
                "value": final_value
            })
#**********************************************************************************************************   
    @frappe.whitelist()
    def get_parameters(self):
        template = frappe.get_doc("Chemical Test Template", self.chemical_test_template)
        data = []
        for row in template.items:
            self.append("chemical_test_details",{
                "parameter": row.parameter
            })
#**********************************************************************************************************   
    @frappe.whitelist()
    def add_data(self):
        readings = [self.row1, self.row2, self.row3, self.row4, self.row5, self.row6]
        valid_readings = [r for r in readings if r not in (None, "", 0)]
        avg = sum(valid_readings) / len(valid_readings) if valid_readings else 0
        for row in self.brinell_hardness_child:
            if (
                row.location == self.location and
                row.test == self.test and
                row.scale == self.scale and
                row.load == self.load and
                row.dial == self.dial and
                row.indentor == self.indentor and
                row.min == self.min and
                row.max == self.max and
                row.avg == avg
            ):
                frappe.throw("Same data already exists in table. No changes detected.")
        self.append("brinell_hardness_child", {
            "location": self.location,
            "test": self.test,
            "scale": self.scale,
            "load": self.load,
            "dial": self.dial,
            "indentor": self.indentor,
            "min": self.min,
            "max": self.max,
            "avg": avg
        })

#**********************************************************************************************************
    @frappe.whitelist()
    def get_avg_absorbed_energy(self):
        total = 0
        count = 0
        for row in self.absorbed_energy_in_joulesj_of_each_specimen:
            if row.range not in (None, "", 0):
                total += row.range
                count += 1
        # self.total_absorbed_energy = total
        self.average_aboserbed_energy_of_one_set = total / count if count else 0
#**********************************************************************************************************
    def on_update_after_submit(self):
        self.db_set("test_completion_date", nowdate())
        self.db_set("test_completion_time", nowtime())
#*******************************************************************************************************
    @frappe.whitelist()
    def update_metallography_rows(self):
        qty = cint(self.no_of_fields or 0)
        existing = len(self.metallography_test_pp or [])
        if qty > existing:
            for _ in range(qty - existing):
                self.append("metallography_test_pp", {})
        elif qty < existing:
            self.set(
                "metallography_test_pp",
                self.metallography_test_pp[:qty]
            )
        return qty
#**************************************************************************************************
    # @frappe.whitelist()
    # def update_metallography_test_table_rows(self):
    #     qty=cint(self.no_of_field or 0)
    #     existing = len(self.metallography_test_table or [])
    #     if qty > existing:
    #         for _ in range(qty - existing):
    #             self.append("metallography_test_table", {})
    #     elif qty < existing:
    #         self.set(
    #             "metallography_test_table",
    #             self.metallography_test_table[:qty]
    #         )
    #     return qty
#**************************************************************************************************