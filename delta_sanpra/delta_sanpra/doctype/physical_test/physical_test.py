# Copyright (c) 2026, Sanpra Software Solution and contributors
# For license information, please see license.txt
import frappe
import re
from frappe.model.document import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import io


class PhysicalTest(Document):
	def before_save(self):
		# Process PDF only if table empty
		if self.pdf_file and not self.test_details_physical:
			self.read_pdf()

		# Process Excel
		if self.excel_file:
			self.upload_excel_file()

	@frappe.whitelist()
	def read_pdf(self):
		# Always clear table first
		# self.set("test_details_physical", [])
		if not self.pdf_file:
			return
		if not self.specimen_shape:
			frappe.throw("Please select Specimen Shape first")
		file_doc = frappe.get_doc("File", {"file_url": self.pdf_file})
		pdf_path = file_doc.get_full_path()
		reader = PdfReader(pdf_path)
		text = ""
		for page in reader.pages:
			page_text = page.extract_text()
			if page_text:
				text += page_text + "\n"
		full_text_lower = text.lower()
		# CHECK Specimen Shape match in PDF
		if self.specimen_shape.lower() not in full_text_lower:
			frappe.throw(f"Specimen Shape '{self.specimen_shape}' not found in attached PDF")
		#Required Parameters based on Shape
		solid_round_params = [
			"Specimen Shape",
			"Initial G.L. For % elong",
			"Specimen Cross Section Area",
			"Final Gauge Length",
			"Final Area",
			"Load at Peak",
			"Specimen Diameter",
			"Final Sp Diameter",
			"Tensile Strength",
			"% Reduction Area",
			"% Elongation",
			"Proof Load 1",
			"Proof Stress 1( 0.2 %)",
			"YS/UTS"
		]
		flat_params = [
			"Specimen Shape",
			"Specimen Width",
			"Specimen Thickness",
			"Initial G.L. For % elong",
			"Specimen Cross Section Area",
			"Final Gauge Length",
			"Load at Peak",
			"Load At Yield",
			"Yield Stress",
			"Tensile Strength",
			"% Elongation",
			"YS/UTS"
		]
		hollow_round_params = [
			"Specimen Shape",
			"Initial G.L. For % elong",
			"Specimen Cross Section Area",
			"Final Gauge Length",
			"Load at Peak",
			"Specimen Inner Diameter",
			"Specimen Outer Diameter",
			"Load At Yield",
			"Yield Stress",
			"Tensile Strength",
			"% Elongation",
			"YS/UTS"
		]
		Odd_Shape_params = [
			"SpecimenType",
			"Initial G.L. For % elong",
			"Specimen Cross Section Area",
			"Final Gauge Length",
			"Load at Peak",
			"Load At Yield",
			"Yield Stress",
			"Tensile Strength",
			"% Elongation",
			"weight/meter",
			"YS/UTS"
		]
		if self.specimen_shape.lower() == "solid round":
			required_params = solid_round_params
		elif self.specimen_shape.lower() == "flat":
			required_params = flat_params
		elif self.specimen_shape.lower() == "hollow round":
			required_params = hollow_round_params
		elif self.specimen_shape.lower() == "odd shape":
			required_params = Odd_Shape_params
		else:
			frappe.throw("Unsupported Specimen Shape")
		required_params_lower = [p.lower().strip() for p in required_params]
		# Extract Only Required Parameters (Case Insensitive)
		# Only keep rows not from PDF (manual or Excel edits stay)
		new_table = []
		added_params = set()

		for row in self.test_details_physical:
			if getattr(row, "from_pdf", 0):
				continue  # skip old PDF rows
			new_table.append(row)
			added_params.add(row.parameter.lower().strip())

		self.set("test_details_physical", new_table)

		for line in text.split("\n"):
			if ":" not in line:
				continue
			param, val = line.split(":", 1)
			param = param.strip()
			val = val.strip()
			normalized_param = param.lower().strip()
			# match ignoring capital/small
			if normalized_param not in required_params_lower:
				continue			# Skip if already added
			if normalized_param in added_params:
				continue
			uom = ""
			if val:
				match = re.match(r"^\s*([-+]?\d*\.?\d+)\s*([^\s]+)", val)
				if match:
					val = match.group(1)
					uom = match.group(2).strip()
					# Remove junk words like Output, Data etc.
					uom = re.split(r'(output|data)', uom, flags=re.IGNORECASE)[0]
					# Keep only valid unit characters
					uom = re.match(r'[a-zA-Z0-9/%²³]+', uom)
					uom = uom.group(0) if uom else ""
			self.append("test_details_physical", {
				"parameter": param,
				"value": val,
				"uom": uom,
    			"from_pdf": 1
			})
			added_params.add(normalized_param)

#**************************************************************************
	@frappe.whitelist()
	def upload_excel_file(self):
		if not self.excel_file or not self.excel_file.lower().endswith(".xlsx"):
			return
		file_name = frappe.db.get_value("File", {"file_url": self.excel_file}, "name")
		if not file_name:
			frappe.throw("Excel file not found in File table")
		file_doc = frappe.get_doc("File", file_name)
		file_content = file_doc.get_content()
		wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
		sheet = wb.active
		self.test_details_physical = [row for row in self.test_details_physical if not getattr(row, "from_excel", 0)]
		existing = {row.parameter: row for row in self.test_details_physical if row.parameter}
		for row in sheet.iter_rows(min_row=2, values_only=True):
			if not row or len(row) < 2:
				continue
			param, val, min_range,max_range,uom = row[0], row[1], row[2] ,row[3],row[4]
			if not param:
				continue
			val = val if val is not None else ""
			min_range = min_range if min_range is not None else ""
			max_range = max_range if max_range is not None else ""
			uom = uom if uom is not None  else ""
			if param in existing:
				existing[param].value = val
				existing[param].min_range = min_range
				existing[param].max_range = max_range
				existing[param].uom = uom
			else:
				self.append("test_details_physical", {
                    "parameter": param, 
                    "value": val, 
                    "min_range": min_range, 
                    "max_range": max_range ,
                    "uom": uom,
                    "from_excel": 1})
#*************************************************************************************************
	@frappe.whitelist()
	def add_data(self):
		readings = [self.row1, self.row2, self.row3, self.row4, self.row5, self.row6]
		valid_readings = [r for r in readings if r not in (None, "", 0)]
		avg = sum(valid_readings) / len(valid_readings) if valid_readings else 0
		for row in self.brinell_hardness_child:
			if (
				row.location == self.location and
				row.test == self.test and
				row.scale == self.scale and
				row.load == self.load and
				row.dial == self.dial and
				row.indentor == self.indentor and
				row.min == self.min and
				row.max == self.max and
				row.r1 == self.row1 and
				row.r2 == self.row2 and
				row.r3 == self.row3 and
				row.r4 == self.row4 and
				row.r5 == self.row5 and
				row.r6 == self.row6 and
				row.avg == avg
			):
				frappe.throw("Same data already exists in table. No changes detected.")
		self.append("brinell_hardness_child", {
			"location": self.location,
			"test": self.test,
			"scale": self.scale,
			"load": self.load,
			"dial": self.dial,
			"indentor": self.indentor,
			"min": self.min,
			"max": self.max,
			"r1": self.row1,
			"r2": self.row2,
			"r3": self.row3,
			"r4": self.row4,
			"r5": self.row5,
			"r6": self.row6,
			"avg": avg
		})
#**************************************************************************
	@frappe.whitelist()
	def set_ulr_counter(self):
		if self.ulr_no:
			return
		company_name = "DELTAA METALLIX SOLUTIONS PRIVATE LIMITED"
		last = frappe.db.get_value("Company", company_name, "custom_ulr_counter") or 0
		count = int(last) + 1
		frappe.db.set_value("Company", company_name, "custom_ulr_counter", count)
		return count

#**************************************************************************
	@frappe.whitelist()
	def get_highlight_colors(self):
		tables = ["test_details_physical"]
		colors = {}
		for table in tables:
			if not hasattr(self, table):
				continue
			for row in getattr(self, table):
				try:
					value = float(row.value) if row.value not in (None, "") else None
					min_range = float(row.min_range) if row.min_range not in (None, "") else None
					max_range = float(row.max_range) if row.max_range not in (None, "") else None

					# Highlight if value < min_range OR value > max_range
					if value is not None and min_range is not None and max_range is not None:
						if value < min_range or value > max_range:
							colors[row.name] = "#2ecc71"     # RED
						else:
							colors[row.name] = "#FF5C5C"     # GREEN
					else:
						colors[row.name] = ""
				except:
					colors[row.name] = ""
		return colors